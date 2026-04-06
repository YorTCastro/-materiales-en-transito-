# app.py — Interfaz web Streamlit con login para Materiales en Tránsito

import io
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth

from loader import load_and_validate
from transformer import aggregate_movements, merge_mb51_with_po
from analyzer import (
    build_detail_df,
    cancellation_rate,
    compute_kpis,
    compute_trend,
    generate_recommendations,
    top_materials_by_avg_time,
    top_pending_amount,
    top_users_overdue,
)
from reporter import (
    write_analysis_sheet,
    write_detail_sheet,
    write_kpis_sheet,
    write_recommendations_sheet,
    write_trend_sheet,
    save_workbook,
)
from config import MB51_COLS, PO_COLS, MB51_KEY_COLS, PO_KEY_COLS, POLICY_DAYS
from openpyxl import Workbook

# ── Configuración de página ───────────────────────────────────────────────────

st.set_page_config(
    page_title="Materiales en Tránsito",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Estilos CSS ───────────────────────────────────────────────────────────────

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .metric-card {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 12px 16px;
        border-left: 4px solid #2F5597;
        margin-bottom: 8px;
    }
    .alert-verde    { background-color: #e8f5e9; color: #1b5e20; padding: 3px 8px; border-radius: 4px; font-weight: bold; }
    .alert-amarillo { background-color: #fff8e1; color: #f57f17; padding: 3px 8px; border-radius: 4px; font-weight: bold; }
    .alert-rojo     { background-color: #ffebee; color: #b71c1c; padding: 3px 8px; border-radius: 4px; font-weight: bold; }
    .login-title { text-align: center; font-size: 2rem; font-weight: bold; color: #2F5597; margin-bottom: 0.5rem; }
    .login-sub   { text-align: center; color: #666; margin-bottom: 2rem; }
</style>
""", unsafe_allow_html=True)


# ── Autenticación ─────────────────────────────────────────────────────────────

# Inicializar claves de sesión requeridas por streamlit-authenticator
for _key in ["authentication_status", "name", "username", "logout"]:
    if _key not in st.session_state:
        st.session_state[_key] = None

with open("credentials.yaml") as f:
    _config = yaml.load(f, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    _config["credentials"],
    _config["cookie"]["name"],
    _config["cookie"]["key"],
    _config["cookie"]["expiry_days"],
)

# Mostrar login centrado cuando no está autenticado
if st.session_state.get("authentication_status") is not True:
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown('<p class="login-title">📦 Materiales en Tránsito</p>', unsafe_allow_html=True)
        st.markdown('<p class="login-sub">Sistema de análisis de ingresos SAP</p>', unsafe_allow_html=True)
        authenticator.login(location="main")

        if st.session_state.get("authentication_status") is False:
            st.error("Usuario o contraseña incorrectos.")
        elif st.session_state.get("authentication_status") is None:
            st.info("Ingresa tus credenciales para continuar.")
    st.stop()


# ── App principal (solo si está autenticado) ──────────────────────────────────

# ── Funciones de plantillas (definidas antes del sidebar) ─────────────────────

@st.cache_data(show_spinner=False)
def generar_plantilla_mb51():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    AZUL="FF2F5597"; BLANCO="FFFFFFFF"; AMARILLO="FFFFF2CC"; VERDE="FFE2EFDA"
    def hc(cell, texto, bg=AZUL):
        cell.value=texto; cell.fill=PatternFill("solid",fgColor=bg)
        cell.font=Font(bold=True,color=BLANCO if bg==AZUL else "FF000000",size=10)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        t=Side(style="thin",color="FF000000"); cell.border=Border(left=t,right=t,top=t,bottom=t)
    def dc(cell,valor,bg=None,italic=False,size=10):
        cell.value=valor
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
        cell.font=Font(size=size,italic=italic,color="FF595959" if italic else "FF000000")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        t=Side(style="thin",color="FFD9D9D9"); cell.border=Border(left=t,right=t,top=t,bottom=t)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="MB51 - Movimientos"
    ws.merge_cells("A1:R1"); ws["A1"].value="PLANTILLA MB51 — Movimientos de Materiales (Exportación SAP)"
    ws["A1"].fill=PatternFill("solid",fgColor=AZUL); ws["A1"].font=Font(bold=True,color=BLANCO,size=12)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=25
    ws.merge_cells("A2:R2"); ws["A2"].value="FORMATO: Exportar desde SAP transacción MB51 como .TXT o .CSV separado por punto y coma (;). Movimiento 101 = Ingreso | 102 = Anulación"
    ws["A2"].fill=PatternFill("solid",fgColor=AMARILLO); ws["A2"].font=Font(color="FF7F6000",size=10,italic=True)
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); ws.row_dimensions[2].height=30
    cols=[("Material","Código material SAP\nEj: 100001234"),("Texto breve de material","Descripción del material"),
          ("CMv","Tipo movimiento\n101=Ingreso\n102=Anulación"),("Alm.","Código almacén\nEj: HB01"),
          ("Pedido","N° pedido compra\nEj: 4500001234"),("Pos.","Posición pedido\nEj: 10"),
          ("Doc.mat.","N° documento material"),("Pos","Posición documento\nEj: 1"),
          ("Cantidad","Cantidad\nComa decimal\nEj: 10,000"),("UMB","Unidad medida\nEj: UN, KG, M"),
          ("Fecha doc.","Fecha documento\nDD.MM.AAAA"),("Reserva","N° reserva\nPuede ir vacío"),
          ("Fe.contab.","Fecha contabilización\nDD.MM.AAAA"),("Hora","Hora movimiento\nEj: 14:30:00"),
          ("Importe ML","Importe moneda local\nEj: 1.500,00"),("Texto cab.documento","Texto cabecera\nPuede ir vacío"),
          ("Referencia","Referencia externa\nPuede ir vacío"),("Usuario","Usuario SAP\nEj: JPEREZ")]
    for i,(n,_) in enumerate(cols,1):
        hc(ws.cell(3,i),n); ws.column_dimensions[get_column_letter(i)].width=16
    ws.row_dimensions[3].height=20
    for i,(_,t) in enumerate(cols,1): dc(ws.cell(4,i),t,bg="FFF2F2F2",italic=True,size=8)
    ws.row_dimensions[4].height=60
    for r,row in enumerate([
        ["100001234","VALVULA ESFERICA 1","101","HB01","4500001234","10","5000001001","1","10,000","UN","01.03.2024","","05.03.2024","10:30:00","1.500,00","","","JPEREZ"],
        ["100001234","VALVULA ESFERICA 1","101","HB01","4500001234","20","5000001002","1","5,000","UN","01.03.2024","","04.03.2024","09:15:00","750,00","","","MGARCIA"],
        ["100002567","TORNILLO HEX M12x50","101","HB02","4500001235","10","5000001003","1","100,000","UN","01.03.2024","","15.03.2024","14:00:00","200,00","","","JPEREZ"],
        ["100002567","TORNILLO HEX M12x50","102","HB02","4500001235","10","5000001004","1","20,000","UN","01.03.2024","","16.03.2024","11:00:00","40,00","ANULACION","","MGARCIA"],
        ["100003891","TUBO ACERO 2 SCH40","101","HB01","4500001236","10","5000001005","1","50,000","M","05.03.2024","","07.03.2024","08:45:00","5.000,00","","","CLOPEZ"],
    ],5):
        bg=VERDE if row[2]=="101" else "FFFFCCCC"
        for c,v in enumerate(row,1): dc(ws.cell(r,c),v,bg=bg)
        ws.row_dimensions[r].height=18
    ws.freeze_panes="A5"; buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


@st.cache_data(show_spinner=False)
def generar_plantilla_pedidos():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    AZUL="FF2F5597"; BLANCO="FFFFFFFF"; AMARILLO="FFFFF2CC"; VERDE="FFE2EFDA"
    def hc(cell,texto,bg=AZUL):
        cell.value=texto; cell.fill=PatternFill("solid",fgColor=bg)
        cell.font=Font(bold=True,color=BLANCO if bg==AZUL else "FF000000",size=10)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        t=Side(style="thin",color="FF000000"); cell.border=Border(left=t,right=t,top=t,bottom=t)
    def dc(cell,valor,bg=None,italic=False,size=10):
        cell.value=valor
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
        cell.font=Font(size=size,italic=italic,color="FF595959" if italic else "FF000000")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        t=Side(style="thin",color="FFD9D9D9"); cell.border=Border(left=t,right=t,top=t,bottom=t)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Pedidos - Doc. Compra"
    ws.merge_cells("A1:J1"); ws["A1"].value="PLANTILLA PEDIDOS — Documentos de Compra (Exportación SAP)"
    ws["A1"].fill=PatternFill("solid",fgColor=AZUL); ws["A1"].font=Font(bold=True,color=BLANCO,size=12)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=25
    ws.merge_cells("A2:J2"); ws["A2"].value="FORMATO: Exportar desde SAP transacciones ME2M / ME2L como .TXT o .CSV separado por punto y coma (;)."
    ws["A2"].fill=PatternFill("solid",fgColor=AMARILLO); ws["A2"].font=Font(color="FF7F6000",size=10,italic=True)
    ws["A2"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); ws.row_dimensions[2].height=25
    cols=[("Material","Código material SAP\nEj: 100001234"),("Texto breve","Descripción del material"),
          ("Cantidad","Cantidad total pedido\nComa decimal\nEj: 15,000"),("Por entrg.","Cantidad pendiente\npor entregar\nEj: 5,000"),
          ("UMP","Unidad de medida\nEj: UN, KG, M"),("Fecha doc.","Fecha del pedido\nDD.MM.AAAA\nEj: 01.03.2024"),
          ("Doc.compr.","N° documento compra\nEj: 4500001234"),("Pos.","Posición pedido\nEj: 10"),
          ("Proveedor/Centro suministrador","Nombre proveedor o\ncentro suministrador\nEj: La Estrella / Taller"),("Mon.","Moneda\nEj: CLP, USD")]
    for i,(n,_) in enumerate(cols,1):
        hc(ws.cell(3,i),n); ws.column_dimensions[get_column_letter(i)].width=20
    ws.row_dimensions[3].height=20
    for i,(_,t) in enumerate(cols,1): dc(ws.cell(4,i),t,bg="FFF2F2F2",italic=True,size=8)
    ws.row_dimensions[4].height=60
    for r,row in enumerate([
        ["100001234","VALVULA ESFERICA 1","15,000","0,000","UN","01.03.2024","4500001234","10","Almacen La Estrella","CLP"],
        ["100001234","VALVULA ESFERICA 1","10,000","0,000","UN","01.03.2024","4500001234","20","Almacen La Estrella","CLP"],
        ["100002567","TORNILLO HEX M12x50","100,000","20,000","UN","01.03.2024","4500001235","10","Taller Central","CLP"],
        ["100003891","TUBO ACERO 2 SCH40","50,000","0,000","M","05.03.2024","4500001236","10","Almacen La Estrella","CLP"],
        ["100004512","BRIDA SLIP-ON 2","30,000","30,000","UN","10.03.2024","4500001237","10","Taller Central","CLP"],
    ],5):
        for c,v in enumerate(row,1): dc(ws.cell(r,c),v,bg=VERDE)
        ws.row_dimensions[r].height=18
    ws.freeze_panes="A5"; buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# Sidebar
with st.sidebar:
    st.markdown(f"### 👤 {st.session_state.get('name', 'Usuario')}")
    st.caption(f"Usuario: `{st.session_state.get('username', '')}`")
    st.divider()
    authenticator.logout("Cerrar sesión", location="sidebar")
    st.divider()

    st.markdown("### ⚙️ Configuración")
    policy_days = st.number_input(
        "Umbral de política (días)",
        min_value=1, max_value=30,
        value=POLICY_DAYS,
        help="Días máximos permitidos entre fecha del pedido y fecha de ingreso."
    )
    st.caption(f"🟢 Verde: 1–3 días  \n🟡 Amarillo: 4–{policy_days} días  \n🔴 Rojo: >{policy_days} días")
    st.divider()
    st.caption("v1.0 · Materiales en Tránsito")

    @st.cache_data(show_spinner=False)
    def generar_plantilla_mb51():
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        AZUL = "FF2F5597"; BLANCO = "FFFFFFFF"
        AMARILLO = "FFFFF2CC"; VERDE = "FFE2EFDA"

        def hc(cell, texto, bg=AZUL):
            cell.value = texto
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.font  = Font(bold=True, color=BLANCO if bg==AZUL else "FF000000", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            t = Side(style="thin", color="FF000000")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

        def dc(cell, valor, bg=None, italic=False, size=10):
            cell.value = valor
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=size, italic=italic, color="FF595959" if italic else "FF000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            t = Side(style="thin", color="FFD9D9D9")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MB51 - Movimientos"

        ws.merge_cells("A1:R1")
        ws["A1"].value = "PLANTILLA MB51 — Movimientos de Materiales (Exportación SAP)"
        ws["A1"].fill  = PatternFill("solid", fgColor=AZUL)
        ws["A1"].font  = Font(bold=True, color=BLANCO, size=12)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws.merge_cells("A2:R2")
        ws["A2"].value = "FORMATO: Exportar desde SAP transacción MB51 como .TXT o .CSV separado por punto y coma (;). Movimiento 101 = Ingreso | 102 = Anulación"
        ws["A2"].fill  = PatternFill("solid", fgColor=AMARILLO)
        ws["A2"].font  = Font(color="FF7F6000", size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        cols = [
            ("Material","Código material SAP\nEj: 100001234"),
            ("Texto breve de material","Descripción del material"),
            ("CMv","Tipo movimiento\n101=Ingreso\n102=Anulación"),
            ("Alm.","Código almacén\nEj: HB01"),
            ("Pedido","N° pedido compra\nEj: 4500001234"),
            ("Pos.","Posición pedido\nEj: 10"),
            ("Doc.mat.","N° documento material"),
            ("Pos","Posición documento\nEj: 1"),
            ("Cantidad","Cantidad\nComa decimal\nEj: 10,000"),
            ("UMB","Unidad medida\nEj: UN, KG, M"),
            ("Fecha doc.","Fecha documento\nDD.MM.AAAA"),
            ("Reserva","N° reserva\nPuede ir vacío"),
            ("Fe.contab.","Fecha contabilización\nDD.MM.AAAA"),
            ("Hora","Hora movimiento\nEj: 14:30:00"),
            ("Importe ML","Importe moneda local\nEj: 1.500,00"),
            ("Texto cab.documento","Texto cabecera\nPuede ir vacío"),
            ("Referencia","Referencia externa\nPuede ir vacío"),
            ("Usuario","Usuario SAP\nEj: JPEREZ"),
        ]
        for i, (n, _) in enumerate(cols, 1):
            hc(ws.cell(3, i), n)
            ws.column_dimensions[get_column_letter(i)].width = 16
        ws.row_dimensions[3].height = 20
        for i, (_, t) in enumerate(cols, 1):
            dc(ws.cell(4, i), t, bg="FFF2F2F2", italic=True, size=8)
        ws.row_dimensions[4].height = 60

        ejemplos = [
            ["100001234","VALVULA ESFERICA 1","101","HB01","4500001234","10","5000001001","1","10,000","UN","01.03.2024","","05.03.2024","10:30:00","1.500,00","","","JPEREZ"],
            ["100001234","VALVULA ESFERICA 1","101","HB01","4500001234","20","5000001002","1","5,000","UN","01.03.2024","","04.03.2024","09:15:00","750,00","","","MGARCIA"],
            ["100002567","TORNILLO HEX M12x50","101","HB02","4500001235","10","5000001003","1","100,000","UN","01.03.2024","","15.03.2024","14:00:00","200,00","","","JPEREZ"],
            ["100002567","TORNILLO HEX M12x50","102","HB02","4500001235","10","5000001004","1","20,000","UN","01.03.2024","","16.03.2024","11:00:00","40,00","ANULACION","","MGARCIA"],
            ["100003891","TUBO ACERO 2 SCH40","101","HB01","4500001236","10","5000001005","1","50,000","M","05.03.2024","","07.03.2024","08:45:00","5.000,00","","","CLOPEZ"],
        ]
        for r, row in enumerate(ejemplos, 5):
            bg = VERDE if row[2] == "101" else "FFFFCCCC"
            for c, v in enumerate(row, 1):
                dc(ws.cell(r, c), v, bg=bg)
            ws.row_dimensions[r].height = 18

        ws.merge_cells("A11:R11")
        ws["A11"].value = "VERDE = Movimiento 101 (Ingreso)     ROJO = Movimiento 102 (Anulación)"
        ws["A11"].fill  = PatternFill("solid", fgColor="FFF2F2F2")
        ws["A11"].font  = Font(italic=True, size=9)
        ws["A11"].alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A5"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @st.cache_data(show_spinner=False)
    def generar_plantilla_pedidos():
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        AZUL = "FF2F5597"; BLANCO = "FFFFFFFF"
        AMARILLO = "FFFFF2CC"; VERDE = "FFE2EFDA"

        def hc(cell, texto, bg=AZUL):
            cell.value = texto
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.font  = Font(bold=True, color=BLANCO if bg==AZUL else "FF000000", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            t = Side(style="thin", color="FF000000")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

        def dc(cell, valor, bg=None, italic=False, size=10):
            cell.value = valor
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=size, italic=italic, color="FF595959" if italic else "FF000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            t = Side(style="thin", color="FFD9D9D9")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedidos - Doc. Compra"

        ws.merge_cells("A1:J1")
        ws["A1"].value = "PLANTILLA PEDIDOS — Documentos de Compra (Exportación SAP)"
        ws["A1"].fill  = PatternFill("solid", fgColor=AZUL)
        ws["A1"].font  = Font(bold=True, color=BLANCO, size=12)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws.merge_cells("A2:J2")
        ws["A2"].value = "FORMATO: Exportar desde SAP transacciones ME2M / ME2L como .TXT o .CSV separado por punto y coma (;)."
        ws["A2"].fill  = PatternFill("solid", fgColor=AMARILLO)
        ws["A2"].font  = Font(color="FF7F6000", size=10, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 25

        cols = [
            ("Material","Código material SAP\nEj: 100001234"),
            ("Texto breve","Descripción del material"),
            ("Cantidad","Cantidad total pedido\nComa decimal\nEj: 15,000"),
            ("Por entrg.","Cantidad pendiente\npor entregar\nEj: 5,000"),
            ("UMP","Unidad de medida\nEj: UN, KG, M"),
            ("Fecha doc.","Fecha del pedido\nDD.MM.AAAA\nEj: 01.03.2024"),
            ("Doc.compr.","N° documento compra\nEj: 4500001234"),
            ("Pos.","Posición pedido\nEj: 10"),
            ("Proveedor/Centro suministrador","Nombre proveedor o\ncentro suministrador\nEj: La Estrella / Taller"),
            ("Mon.","Moneda\nEj: CLP, USD"),
        ]
        for i, (n, _) in enumerate(cols, 1):
            hc(ws.cell(3, i), n)
            ws.column_dimensions[get_column_letter(i)].width = 20
        ws.row_dimensions[3].height = 20
        for i, (_, t) in enumerate(cols, 1):
            dc(ws.cell(4, i), t, bg="FFF2F2F2", italic=True, size=8)
        ws.row_dimensions[4].height = 60

        ejemplos = [
            ["100001234","VALVULA ESFERICA 1","15,000","0,000","UN","01.03.2024","4500001234","10","Almacen La Estrella","CLP"],
            ["100001234","VALVULA ESFERICA 1","10,000","0,000","UN","01.03.2024","4500001234","20","Almacen La Estrella","CLP"],
            ["100002567","TORNILLO HEX M12x50","100,000","20,000","UN","01.03.2024","4500001235","10","Taller Central","CLP"],
            ["100003891","TUBO ACERO 2 SCH40","50,000","0,000","M","05.03.2024","4500001236","10","Almacen La Estrella","CLP"],
            ["100004512","BRIDA SLIP-ON 2","30,000","30,000","UN","10.03.2024","4500001237","10","Taller Central","CLP"],
        ]
        for r, row in enumerate(ejemplos, 5):
            for c, v in enumerate(row, 1):
                dc(ws.cell(r, c), v, bg=VERDE)
            ws.row_dimensions[r].height = 18

        ws.merge_cells("A11:J11")
        ws["A11"].value = "El campo Proveedor/Centro suministrador debe contener 'La Estrella' o 'Taller' para identificar el origen automaticamente."
        ws["A11"].fill  = PatternFill("solid", fgColor=AMARILLO)
        ws["A11"].font  = Font(italic=True, size=9, color="FF7F6000")
        ws["A11"].alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A5"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    st.divider()
    st.caption("v1.0 · Materiales en Tránsito")

# Título principal
st.title("📦 Análisis de Materiales en Tránsito")

# ── Carga de archivos + Plantillas ────────────────────────────────────────────

col_a, col_b = st.columns(2)
with col_a:
    mb51_file = st.file_uploader(
        "📂 Archivo MB51 (Movimientos de materiales)",
        type=["csv", "txt", "xlsx"],
        help="Exportación de la transacción MB51 de SAP."
    )
    st.download_button(
        label="📄 Descargar plantilla MB51",
        data=generar_plantilla_mb51(),
        file_name="Plantilla_MB51.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with col_b:
    po_file = st.file_uploader(
        "📂 Archivo de Pedidos (Documentos de compra)",
        type=["csv", "txt", "xlsx"],
        help="Exportación del listado de pedidos/documentos de compra de SAP."
    )
    st.download_button(
        label="📄 Descargar plantilla Pedidos",
        data=generar_plantilla_pedidos(),
        file_name="Plantilla_Pedidos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

if not mb51_file or not po_file:
    st.info("⬆️ Sube los dos archivos para comenzar el análisis.")
    st.stop()

# ── Procesamiento ─────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def process_files(mb51_bytes, mb51_name, po_bytes, po_name, policy_days):
    """Procesa los archivos y retorna todos los DataFrames y KPIs."""

    def save_temp(content, name):
        suffix = Path(name).suffix or ".csv"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(content)
        tmp.flush()
        return tmp.name

    mb51_path = save_temp(mb51_bytes, mb51_name)
    po_path   = save_temp(po_bytes, po_name)

    mb51_df, mb51_col_map = load_and_validate(
        mb51_path, MB51_COLS, MB51_KEY_COLS, "MB51", interactive=False
    )
    po_df, po_col_map = load_and_validate(
        po_path, PO_COLS, PO_KEY_COLS, "Pedidos", interactive=False
    )

    agg_mb51  = aggregate_movements(mb51_df, mb51_col_map)
    merged    = merge_mb51_with_po(agg_mb51, po_df, po_col_map)
    detail    = build_detail_df(merged, policy_days=policy_days)
    kpis      = compute_kpis(detail)
    top_mat   = top_materials_by_avg_time(detail)
    top_users = top_users_overdue(detail)
    weekly, monthly = compute_trend(detail)
    top_pend  = top_pending_amount(detail)
    cancel_df = cancellation_rate(detail)
    recs      = generate_recommendations(kpis, detail, top_mat, top_users, policy_days)

    return detail, kpis, top_mat, top_users, weekly, monthly, top_pend, cancel_df, recs

with st.spinner("Procesando archivos SAP..."):
    try:
        detail, kpis, top_mat, top_users, weekly, monthly, top_pend, cancel_df, recs = process_files(
            mb51_file.read(), mb51_file.name,
            po_file.read(), po_file.name,
            policy_days,
        )
    except Exception as e:
        st.error(f"Error al procesar los archivos: {e}")
        st.caption("Verifica que los archivos sean exportaciones válidas de SAP con las columnas correctas.")
        st.stop()

# ── Paleta de colores profesional ────────────────────────────────────────────
import plotly.graph_objects as go
import plotly.express as px

C_NAVY    = "#1B2A4A"   # Azul marino oscuro
C_BLUE    = "#2563EB"   # Azul corporativo
C_TEAL    = "#0D9488"   # Verde azulado
C_GREEN   = "#16A34A"   # Verde éxito
C_AMBER   = "#D97706"   # Ámbar advertencia
C_RED     = "#DC2626"   # Rojo alerta
C_SLATE   = "#64748B"   # Gris pizarra
C_LIGHT   = "#F8FAFC"   # Fondo claro

CHART_FONT = dict(family="Inter, Arial, sans-serif", size=12, color="#374151")
CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=CHART_FONT,
    margin=dict(t=50, b=20, l=10, r=20),
    hoverlabel=dict(bgcolor="white", font_size=12, bordercolor="#e5e7eb"),
)

# ── KPIs ──────────────────────────────────────────────────────────────────────

st.divider()
st.subheader("📊 Resumen Ejecutivo")

def kpi_card(label, value, sublabel, color, icon):
    return f"""
    <div style="background:white; border-radius:10px; padding:16px 20px;
                box-shadow:0 1px 3px rgba(0,0,0,0.08), 0 1px 2px rgba(0,0,0,0.06);
                border-top:3px solid {color}; min-height:95px;">
      <div style="display:flex; justify-content:space-between; align-items:flex-start;">
        <div style="font-size:11px; font-weight:600; color:{C_SLATE};
                    text-transform:uppercase; letter-spacing:0.6px;">{label}</div>
        <div style="font-size:18px; opacity:0.7;">{icon}</div>
      </div>
      <div style="font-size:30px; font-weight:700; color:{color};
                  line-height:1.1; margin:4px 0;">{value}</div>
      <div style="font-size:11px; color:{C_SLATE};">{sublabel}</div>
    </div>"""

c1,c2,c3,c4,c5 = st.columns(5)
c1.markdown(kpi_card("Total líneas",    kpis["total_lineas"],                f"{kpis['n_con_entrada']} con entrada",          C_NAVY,  "📋"), unsafe_allow_html=True)
c2.markdown(kpi_card("Sin entrada",     kpis["n_sin_entrada"],               "pendientes de registrar",                        C_SLATE, "⏳"), unsafe_allow_html=True)
c3.markdown(kpi_card("% Oportuno",      f"{kpis['pct_oportuno']}%",          f"{kpis['n_verde']+kpis['n_amarillo']} registros",C_GREEN, "✅"), unsafe_allow_html=True)
c4.markdown(kpi_card("% Fuera política",f"{kpis['pct_vencido']}%",           f"{kpis['n_rojo']} registros vencidos",           C_RED,   "🚨"), unsafe_allow_html=True)
c5.markdown(kpi_card("Imp. Pendiente",  f"{kpis['importe_pendiente_total']:,.0f}", "moneda local",                             C_AMBER, "💰"), unsafe_allow_html=True)

st.markdown("<div style='margin-top:20px'></div>", unsafe_allow_html=True)

# ── Fila de gráficas — 3 columnas ────────────────────────────────────────────
gc1, gc2, gc3 = st.columns(3)

# Gráfica 1: Donut — estado de ingresos
with gc1:
    labels = ["Oportuno", "En límite", "Vencido", "Sin entrada"]
    values = [kpis["n_verde"], kpis["n_amarillo"], kpis["n_rojo"], kpis["n_sin_entrada"]]
    colors_donut = [C_GREEN, C_AMBER, C_RED, "#94A3B8"]

    fig_donut = go.Figure(go.Pie(
        labels=labels, values=values, hole=0.62,
        marker=dict(colors=colors_donut, line=dict(color="white", width=2.5)),
        textinfo="label+percent",
        textposition="outside",
        textfont=dict(size=11),
        hovertemplate="<b>%{label}</b><br>%{value} registros (%{percent})<extra></extra>",
        direction="clockwise", sort=False,
        showlegend=False,
    ))
    fig_donut.add_annotation(
        text=f"<b>{kpis['total_lineas']}</b><br>líneas",
        x=0.5, y=0.5, showarrow=False,
        font=dict(size=16, color=C_NAVY, family="Arial"),
    )
    fig_donut.update_layout(
        **CHART_LAYOUT,
        title=dict(text="Estado de Ingresos", font=dict(size=13, color=C_NAVY),
                   x=0.5, xanchor="center", y=0.97),
        height=310,
        margin=dict(t=40, b=40, l=60, r=60),
    )
    st.plotly_chart(fig_donut, use_container_width=True)

# Gráfica 2: Barras horizontales — % oportuno y vencido por origen
with gc2:
    origen_kpis = kpis.get("por_origen", {})
    valid_origins = {k: v for k, v in origen_kpis.items() if v is not None}

    if valid_origins:
        names = list(valid_origins.keys())
        pct_op = [round(v, 1) for v in valid_origins.values()]
        pct_venc = [round(100 - v, 1) for v in valid_origins.values()]

        fig_orig = go.Figure()
        fig_orig.add_trace(go.Bar(
            name="Oportuno", y=names, x=pct_op, orientation="h",
            marker=dict(color=C_GREEN, opacity=0.9),
            text=[f"{v}%" for v in pct_op], textposition="inside",
            textfont=dict(color="white", size=12, family="Arial Bold"),
            hovertemplate="<b>%{y}</b><br>Oportuno: %{x}%<extra></extra>",
        ))
        fig_orig.add_trace(go.Bar(
            name="Vencido", y=names, x=pct_venc, orientation="h",
            marker=dict(color=C_RED, opacity=0.75),
            text=[f"{v}%" for v in pct_venc], textposition="inside",
            textfont=dict(color="white", size=12),
            hovertemplate="<b>%{y}</b><br>Vencido: %{x}%<extra></extra>",
        ))
        fig_orig.update_layout(
            **CHART_LAYOUT,
            title=dict(text="Cumplimiento por Origen", font=dict(size=13, color=C_NAVY),
                       x=0.5, xanchor="center", y=0.97),
            barmode="stack",
            xaxis=dict(range=[0, 100], ticksuffix="%", gridcolor="#F1F5F9", tickfont=dict(size=10)),
            yaxis=dict(tickfont=dict(size=11)),
            legend=dict(orientation="h", y=-0.18, x=0.5, xanchor="center", font=dict(size=10)),
            height=310,
            margin=dict(t=40, b=50, l=10, r=10),
        )
        st.plotly_chart(fig_orig, use_container_width=True)
    else:
        st.info("Sin datos de origen para mostrar.")

# Gráfica 3: Indicadores de gestión — barras verticales
with gc3:
    indicadores = ["% Oportuno", "% Vencido", "% Parciales", "% Anulaciones"]
    valores = [kpis["pct_oportuno"], kpis["pct_vencido"],
               kpis["pct_parciales"], kpis["pct_anulaciones"]]
    colores = [C_GREEN, C_RED, C_BLUE, C_AMBER]

    fig_ind = go.Figure()
    for label, val, color in zip(indicadores, valores, colores):
        fig_ind.add_trace(go.Bar(
            x=[label], y=[val],
            name=label,
            marker=dict(color=color, opacity=0.88, cornerradius=5),
            text=[f"<b>{val}%</b>"],
            textposition="outside",
            textfont=dict(size=12),
            hovertemplate=f"<b>{label}</b>: {val}%<extra></extra>",
            width=0.55,
        ))
    fig_ind.add_hline(
        y=80, line=dict(color=C_SLATE, dash="dot", width=1.2),
        annotation=dict(text="Meta", font=dict(size=9, color=C_SLATE)),
        annotation_position="top right",
    )
    fig_ind.update_layout(
        **CHART_LAYOUT,
        title=dict(text="Indicadores de Gestión", font=dict(size=13, color=C_NAVY),
                   x=0.5, xanchor="center", y=0.97),
        yaxis=dict(range=[0, 115], ticksuffix="%", gridcolor="#F1F5F9",
                   zeroline=False, tickfont=dict(size=10)),
        xaxis=dict(tickfont=dict(size=10)),
        showlegend=False,
        height=310,
        margin=dict(t=40, b=20, l=10, r=10),
    )
    st.plotly_chart(fig_ind, use_container_width=True)

# ── Tabs de análisis ──────────────────────────────────────────────────────────

st.divider()
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📋 Detalle por línea",
    "📈 Análisis adicionales",
    "📅 Tendencia",
    "💡 Recomendaciones",
    "⬇️ Descargar Excel",
])

# -- Tab 1: Detalle --
with tab1:
    st.caption("Ordenado por días transcurridos (mayor a menor). La columna Alerta indica el estado de cada ingreso.")

    display_cols = {
        "Material":      "_material_po",
        "Descripción":   "_description",
        "Origen":        "origen",
        "Pedido":        "_order",
        "Pos.":          "_position",
        "Fecha Pedido":  "_po_doc_date",
        "Fecha Ingreso": "last_acc_date",
        "Días":          "dias_transcurridos",
        "Alerta":        "alerta",
        "Qty Ordenada":  "_qty_ordered",
        "Qty Entrada":   "qty_101",
        "Qty Anulada":   "qty_102",
        "Qty Neta":      "qty_neta",
        "Qty Pendiente": "qty_pendiente",
        "Importe ML":    "amount_101",
        "Usuarios":      "users_101",
        "¿Anulación?":   "tiene_anulacion",
    }

    df_show = detail[[c for c in display_cols.values() if c in detail.columns]].copy()
    df_show.columns = [k for k, v in display_cols.items() if v in detail.columns]
    df_show = df_show.sort_values("Días", ascending=False, na_position="last")

    # Solo colorear la columna Alerta — filas limpias sin fondo rosado
    def color_alerta_col(val):
        m = {
            "VERDE":       f"background-color:{C_GREEN}; color:white; font-weight:600; border-radius:4px",
            "AMARILLO":    f"background-color:{C_AMBER}; color:white; font-weight:600; border-radius:4px",
            "ROJO":        f"background-color:{C_RED};   color:white; font-weight:600; border-radius:4px",
            "SIN ENTRADA": f"background-color:{C_SLATE}; color:white; font-weight:600; border-radius:4px",
        }
        return m.get(str(val), "")

    styled = df_show.style.map(color_alerta_col, subset=["Alerta"])
    st.dataframe(styled, use_container_width=True, height=500)

# -- Tab 2: Análisis adicionales --
with tab2:
    r1c1, r1c2 = st.columns(2)

    with r1c1:
        st.markdown("##### 🏆 Top materiales — tiempo promedio de ingreso")
        if not top_mat.empty:
            df_mat = top_mat.sort_values("Días Promedio")
            fig_mat = go.Figure(go.Bar(
                x=df_mat["Días Promedio"], y=df_mat["Material"],
                orientation="h",
                marker=dict(
                    color=df_mat["Días Promedio"],
                    colorscale=[[0, C_TEAL], [0.5, C_AMBER], [1, C_RED]],
                    colorbar=dict(title="Días", thickness=10),
                ),
                text=df_mat["Días Promedio"].apply(lambda v: f"{v:.1f} d"),
                textposition="outside", textfont=dict(size=11),
                hovertemplate="<b>%{y}</b><br>Promedio: %{x:.1f} días<extra></extra>",
            ))
            fig_mat.update_layout(
                **CHART_LAYOUT, height=340,
                xaxis=dict(title="Días promedio", gridcolor="#F1F5F9"),
                yaxis=dict(tickfont=dict(size=10)),
            )
            st.plotly_chart(fig_mat, use_container_width=True)
        else:
            st.caption("Sin datos.")

    with r1c2:
        st.markdown("##### 👤 Usuarios con más ingresos vencidos")
        if not top_users.empty:
            df_usr = top_users.sort_values("Ingresos Vencidos")
            fig_usr = go.Figure(go.Bar(
                x=df_usr["Ingresos Vencidos"], y=df_usr["Usuario"],
                orientation="h",
                marker=dict(color=C_RED, opacity=0.85, cornerradius=4),
                text=df_usr["Ingresos Vencidos"],
                textposition="outside", textfont=dict(size=11),
                hovertemplate="<b>%{y}</b><br>Vencidos: %{x}<extra></extra>",
            ))
            fig_usr.update_layout(
                **CHART_LAYOUT, height=340,
                xaxis=dict(title="Ingresos vencidos", gridcolor="#F1F5F9"),
                yaxis=dict(tickfont=dict(size=11)),
            )
            st.plotly_chart(fig_usr, use_container_width=True)
        else:
            st.caption("Sin datos.")

    r2c1, r2c2 = st.columns(2)

    with r2c1:
        st.markdown("##### 💰 Mayor importe pendiente por material")
        if not top_pend.empty:
            df_p = top_pend.sort_values("Importe Pendiente")
            fig_pend = go.Figure(go.Bar(
                x=df_p["Importe Pendiente"], y=df_p["Material"],
                orientation="h",
                marker=dict(color=C_BLUE, opacity=0.85, cornerradius=4),
                text=df_p["Importe Pendiente"].apply(lambda v: f"{v:,.0f}"),
                textposition="outside", textfont=dict(size=11),
                hovertemplate="<b>%{y}</b><br>Pendiente: %{x:,.0f}<extra></extra>",
            ))
            fig_pend.update_layout(
                **CHART_LAYOUT, height=340,
                xaxis=dict(title="Importe pendiente", gridcolor="#F1F5F9"),
                yaxis=dict(tickfont=dict(size=10)),
            )
            st.plotly_chart(fig_pend, use_container_width=True)
        else:
            st.caption("Sin datos.")

    with r2c2:
        st.markdown("##### 🔄 Tasa de anulaciones por origen y usuario")
        if not cancel_df.empty:
            def color_cancel(val):
                if isinstance(val, (int, float)):
                    if val >= 50: return f"background-color:{C_RED};   color:white; font-weight:600"
                    if val >= 20: return f"background-color:{C_AMBER}; color:white; font-weight:600"
                    if val >  0:  return f"background-color:{C_TEAL};  color:white"
                return ""
            st.dataframe(
                cancel_df.style.map(color_cancel, subset=["% Anulación"]),
                use_container_width=True, hide_index=True, height=340,
            )
        else:
            st.caption("Sin datos.")

# -- Tab 3: Tendencia --
with tab3:
    col_w, col_m = st.columns(2)

    def trend_chart(df_trend, title):
        if df_trend.empty:
            st.caption("Sin suficientes datos para calcular tendencia.")
            return
        df_trend = df_trend.copy()
        df_trend["Período"] = df_trend["Período"].astype(str).str[:10]

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_trend["Período"], y=df_trend["% Oportuno"],
            mode="lines+markers",
            line=dict(color=C_BLUE, width=2.5, shape="spline"),
            marker=dict(size=8, color="white", line=dict(color=C_BLUE, width=2.5)),
            fill="tozeroy", fillcolor=f"rgba(37,99,235,0.07)",
            hovertemplate="<b>%{x}</b><br>% Oportuno: %{y:.1f}%<extra></extra>",
        ))
        # Puntos coloreados por umbral
        for _, row in df_trend.iterrows():
            c = C_GREEN if row["% Oportuno"] >= 80 else C_AMBER if row["% Oportuno"] >= 60 else C_RED
            fig.add_trace(go.Scatter(
                x=[row["Período"]], y=[row["% Oportuno"]],
                mode="markers+text",
                marker=dict(size=10, color=c),
                text=[f"{row['% Oportuno']:.0f}%"],
                textposition="top center", textfont=dict(size=10),
                showlegend=False,
                hoverinfo="skip",
            ))
        fig.add_hline(
            y=80, line=dict(color=C_SLATE, dash="dot", width=1.5),
            annotation=dict(text="Meta 80%", font=dict(size=10, color=C_SLATE),
                            bgcolor="white", borderpad=3),
            annotation_position="top right",
        )
        fig.update_layout(
            **CHART_LAYOUT,
            title=dict(text=title, font=dict(size=14, color=C_NAVY), x=0.5, xanchor="center"),
            yaxis=dict(range=[0, 110], ticksuffix="%", gridcolor="#F1F5F9",
                       zeroline=False, title="% Oportuno"),
            xaxis=dict(tickangle=-30, gridcolor="#F1F5F9"),
            height=340, showlegend=False,
        )
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(df_trend, use_container_width=True, hide_index=True)

    with col_w:
        trend_chart(weekly,  "Tendencia Semanal")
    with col_m:
        trend_chart(monthly, "Tendencia Mensual")

# -- Tab 4: Recomendaciones --
with tab4:
    if recs:
        for i, rec in enumerate(recs, 1):
            is_alta = "PRIORIDAD ALTA" in rec
            color = C_RED if is_alta else C_AMBER if i <= 2 else C_BLUE
            icon  = "🚨" if is_alta else "⚠️" if i <= 2 else "💡"
            label = "Prioridad Alta" if is_alta else "Atención" if i <= 2 else "Sugerencia"
            st.markdown(f"""
            <div style="background:white; border-left:4px solid {color};
                        border-radius:8px; padding:16px 20px; margin-bottom:12px;
                        box-shadow:0 1px 3px rgba(0,0,0,0.07);">
              <div style="display:flex; align-items:center; gap:8px; margin-bottom:6px;">
                <span style="font-size:16px;">{icon}</span>
                <span style="font-size:11px; font-weight:700; color:{color};
                             text-transform:uppercase; letter-spacing:0.5px;">{label} · #{i}</span>
              </div>
              <p style="margin:0; color:#374151; line-height:1.6; font-size:14px;">{rec}</p>
            </div>""", unsafe_allow_html=True)
    else:
        st.success("✅ No se detectaron problemas críticos que recomienden acción inmediata.")

# -- Tab 5: Descargar Excel --
with tab5:
    st.markdown("Genera el informe completo en Excel con todas las hojas y formatos.")

    @st.cache_data(show_spinner=False)
    def generate_excel(detail_bytes, kpis_str, policy_days):
        """Genera el Excel en memoria y retorna los bytes."""
        import pickle, hashlib
        # Reconstruir objetos desde caché
        detail_local = pd.read_parquet(io.BytesIO(detail_bytes))

        kpis_local      = compute_kpis(detail_local)
        top_mat_local   = top_materials_by_avg_time(detail_local)
        top_users_local = top_users_overdue(detail_local)
        weekly_l, monthly_l = compute_trend(detail_local)
        top_pend_local  = top_pending_amount(detail_local)
        cancel_local    = cancellation_rate(detail_local)
        recs_local      = generate_recommendations(kpis_local, detail_local,
                                                    top_mat_local, top_users_local, policy_days)

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        write_kpis_sheet(wb, kpis_local, policy_days)
        write_detail_sheet(wb, detail_local)
        write_analysis_sheet(wb, top_mat_local,   "top_materials", "Top 10 Materiales")
        write_analysis_sheet(wb, top_users_local, "top_users",     "Top 5 Usuarios Demora")
        write_trend_sheet(wb, weekly_l, monthly_l)
        write_analysis_sheet(wb, top_pend_local,  "pending",       "Mayor Importe Pendiente")
        write_analysis_sheet(wb, cancel_local,    "cancellations", "Tasa de Anulaciones")
        write_recommendations_sheet(wb, recs_local)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Serializar detail para caché
    detail_buf = io.BytesIO()
    detail.to_parquet(detail_buf)

    if st.button("📥 Generar Excel", type="primary"):
        with st.spinner("Generando Excel..."):
            try:
                excel_bytes = generate_excel(detail_buf.getvalue(), str(kpis), policy_days)
                from datetime import date
                filename = f"Informe_Transito_{date.today().strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="⬇️ Descargar informe Excel",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
                st.success(f"Excel listo: {filename}")
            except Exception as e:
                st.error(f"Error generando Excel: {e}")
