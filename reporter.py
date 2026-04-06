# reporter.py — Generación del Excel con openpyxl (estilos directos, gráficos)

from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    ALERT_COLOR_MAP,
    COLOR_ALT_ROW,
    COLOR_HEADER,
    COLOR_HEADER_FONT,
    COLOR_HEADER_ORANGE,
    COLOR_SUBHEADER,
    COLOR_WARNING_ROW,
    MAX_COL_WIDTH,
    OUTPUT_SHEETS,
)


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color="FF000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size)


def _border() -> Border:
    thin = Side(style="thin", color="FFD9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_cell(cell, bg_color=None, font_color=COLOR_HEADER_FONT, bold=True, size=11):
    bg = bg_color or COLOR_HEADER
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, color=font_color, size=size)
    cell.alignment = _center()
    cell.border = _border()


def _style_data_cell(cell, bg_color=None, bold=False, number_format=None, align="left"):
    if bg_color:
        cell.fill = _fill(bg_color)
    cell.font = _font(bold=bold)
    cell.alignment = _left() if align == "left" else _center()
    cell.border = _border()
    if number_format:
        cell.number_format = number_format


def auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, MAX_COL_WIDTH)


# ── Hoja 1: Resumen KPIs ──────────────────────────────────────────────────────

def write_kpis_sheet(wb: Workbook, kpis: dict, policy_days: int):
    ws = wb.create_sheet(OUTPUT_SHEETS["kpis"])
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "INFORME DE MATERIALES EN TRÁNSITO"
    c.fill = _fill(COLOR_HEADER)
    c.font = _font(bold=True, color=COLOR_HEADER_FONT, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Subtítulo
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Umbral de política: {policy_days} días"
    c.font = _font(bold=False, color="FF595959", size=10)
    c.alignment = _center()

    ws.append([])  # fila 3 vacía

    # Cabecera de tabla KPIs
    headers = ["Indicador", "Valor", "", ""]
    ws.append(headers)
    for cell in ws[4]:
        _style_header_cell(cell)

    kpi_rows = [
        ("Total líneas de pedido analizadas",    kpis["total_lineas"],           "", ""),
        ("Líneas con entrada registrada (101)",   kpis["n_con_entrada"],          "", ""),
        ("Líneas sin entrada",                    kpis["n_sin_entrada"],          "", ""),
        ("",                                      "",                             "", ""),
        ("Ingresos oportunos (VERDE + AMARILLO)", f"{kpis['pct_oportuno']}%",     "", ""),
        ("  — VERDE (1–3 días)",                  kpis["n_verde"],                "", ""),
        (f"  — AMARILLO (4–{policy_days} días)",  kpis["n_amarillo"],             "", ""),
        ("Ingresos fuera de política (ROJO)",     f"{kpis['pct_vencido']}%",      "", ""),
        ("  — Cantidad registros ROJO",           kpis["n_rojo"],                 "", ""),
        ("",                                      "",                             "", ""),
        ("Líneas con anulación (mov. 102)",       f"{kpis['pct_anulaciones']}%",  "", ""),
        ("  — Cantidad",                          kpis["n_anulaciones"],          "", ""),
        ("Ingresos parciales",                    f"{kpis['pct_parciales']}%",    "", ""),
        ("  — Cantidad",                          kpis["n_parciales"],            "", ""),
        ("",                                      "",                             "", ""),
        ("Importe total pendiente de ingresar",   kpis["importe_pendiente_total"],"", ""),
    ]

    alt = False
    for row_data in kpi_rows:
        ws.append(row_data)
        row_idx = ws.max_row
        if row_data[0] == "":
            continue
        bg = COLOR_ALT_ROW if alt else None
        for col_idx, cell in enumerate(ws[row_idx], 1):
            _style_data_cell(cell, bg_color=bg)
            if col_idx == 2:
                # Resaltar el valor si es % de vencidos alto
                if "pct_vencido" in str(row_data[0]).lower() or "fuera de política" in str(row_data[0]).lower():
                    pct_val = kpis["pct_vencido"]
                    if pct_val > 30:
                        cell.fill = _fill("FFFF0000")
                        cell.font = _font(bold=True, color=COLOR_HEADER_FONT)
                    elif pct_val > 15:
                        cell.fill = _fill("FFFFC000")
                        cell.font = _font(bold=True)
                # Formato de moneda para importe
                if "importe" in str(row_data[0]).lower():
                    cell.number_format = "#,##0.00"
        alt = not alt

    ws.append([])

    # Comparativo por origen
    origen_kpis = kpis.get("por_origen", {})
    if origen_kpis:
        ws.append(["Comparativo por Origen", "% Oportuno", "", ""])
        for cell in ws[ws.max_row]:
            _style_header_cell(cell, bg_color=COLOR_SUBHEADER, font_color="FF000000")

        for origin_name, pct in sorted(origen_kpis.items(),
                                        key=lambda x: (x[1] or 0), reverse=True):
            pct_str = f"{round(pct, 1)}%" if pct is not None else "Sin datos"
            ws.append([origin_name, pct_str, "", ""])
            for cell in ws[ws.max_row]:
                _style_data_cell(cell)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18


# ── Hoja 2: Detalle por línea ─────────────────────────────────────────────────

DETAIL_COLUMNS = [
    ("Material",           "_material_po",       None,         "left"),
    ("Descripción",        "_description",       None,         "left"),
    ("Origen",             "origen",             None,         "center"),
    ("Pedido",             "_order",             None,         "center"),
    ("Posición",           "_position",          None,         "center"),
    ("Fecha Pedido",       "_po_doc_date",       "DD/MM/YYYY", "center"),
    ("Fecha Ingreso",      "last_acc_date",      "DD/MM/YYYY", "center"),
    ("Días Transcurridos", "dias_transcurridos", "0",          "center"),
    ("Alerta",             "alerta",             None,         "center"),
    ("Qty Ordenada",       "_qty_ordered",       "#,##0.00",   "right"),
    ("Qty Entradas (101)", "qty_101",            "#,##0.00",   "right"),
    ("Qty Anulada (102)",  "qty_102",            "#,##0.00",   "right"),
    ("Qty Neta",           "qty_neta",           "#,##0.00",   "right"),
    ("Qty Pendiente",      "qty_pendiente",      "#,##0.00",   "right"),
    ("Importe ML",         "amount_101",         "#,##0.00",   "right"),
    ("Imp. Pendiente Est.","importe_pendiente",  "#,##0.00",   "right"),
    ("Usuarios",           "users_101",          None,         "left"),
    ("¿Anulación?",        "tiene_anulacion",    None,         "center"),
    ("Advertencia",        "_advertencia",       None,         "left"),
]


def write_detail_sheet(wb: Workbook, detail_df: pd.DataFrame):
    ws = wb.create_sheet(OUTPUT_SHEETS["detail"])
    ws.sheet_view.showGridLines = False

    # Cabecera
    headers = [col[0] for col in DETAIL_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        _style_header_cell(cell)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # Ordenar por días descendente (vencidos primero)
    df = detail_df.copy()
    df = df.sort_values("dias_transcurridos", ascending=False, na_position="last")

    for _, row in df.iterrows():
        row_values = []
        for _, src_col, _, _ in DETAIL_COLUMNS:
            val = row.get(src_col, "")
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            row_values.append(val)
        ws.append(row_values)

        row_idx = ws.max_row
        alert_val = str(row.get("alerta", ""))
        warn_val  = str(row.get("_advertencia", ""))
        row_bg = COLOR_WARNING_ROW if warn_val else None

        for col_idx, (_, _, num_fmt, align) in enumerate(DETAIL_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = DETAIL_COLUMNS[col_idx - 1][1]

            # Color de la celda de alerta
            if col_name == "alerta":
                alert_color = ALERT_COLOR_MAP.get(alert_val)
                _style_data_cell(cell, bg_color=alert_color, bold=True,
                                 number_format=num_fmt, align=align)
            elif row_bg:
                _style_data_cell(cell, bg_color=row_bg, number_format=num_fmt, align=align)
            else:
                _style_data_cell(cell, number_format=num_fmt, align=align)

    auto_column_width(ws)


# ── Escritura genérica de una hoja de análisis ───────────────────────────────

def write_analysis_sheet(wb: Workbook, df: pd.DataFrame, sheet_key: str,
                          title: str, header_bg=None):
    ws = wb.create_sheet(OUTPUT_SHEETS[sheet_key])
    ws.sheet_view.showGridLines = False

    # Título
    n_cols = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).fill = _fill(header_bg or COLOR_HEADER)
    ws.cell(1, 1).font = _font(bold=True, color=COLOR_HEADER_FONT, size=12)
    ws.cell(1, 1).alignment = _center()
    ws.row_dimensions[1].height = 25

    if df.empty:
        ws.cell(2, 1).value = "Sin datos para mostrar."
        return

    # Cabecera de columnas
    ws.append(list(df.columns))
    for cell in ws[2]:
        _style_header_cell(cell, bg_color=COLOR_SUBHEADER, font_color="FF000000")

    # Datos
    alt = False
    for _, row in df.iterrows():
        ws.append(list(row))
        bg = COLOR_ALT_ROW if alt else None
        for cell in ws[ws.max_row]:
            _style_data_cell(cell, bg_color=bg)
        alt = not alt

    auto_column_width(ws)


# ── Hoja de tendencia con gráfico ────────────────────────────────────────────

def write_trend_sheet(wb: Workbook, weekly_df: pd.DataFrame, monthly_df: pd.DataFrame):
    ws = wb.create_sheet(OUTPUT_SHEETS["trend"])
    ws.sheet_view.showGridLines = False

    def write_table(start_row, title, df_trend):
        n_cols = max(len(df_trend.columns), 1)
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row, end_column=n_cols
        )
        ws.cell(start_row, 1).value = title
        ws.cell(start_row, 1).fill = _fill(COLOR_HEADER)
        ws.cell(start_row, 1).font = _font(bold=True, color=COLOR_HEADER_FONT, size=12)
        ws.cell(start_row, 1).alignment = _center()
        ws.row_dimensions[start_row].height = 22

        if df_trend.empty:
            ws.cell(start_row + 1, 1).value = "Sin datos."
            return start_row + 3

        # Cabecera
        header_row = start_row + 1
        ws.append(list(df_trend.columns))
        for cell in ws[header_row]:
            _style_header_cell(cell, bg_color=COLOR_SUBHEADER, font_color="FF000000")

        # Datos
        for _, row in df_trend.iterrows():
            row_vals = list(row)
            # Formatear fecha
            if hasattr(row_vals[0], "strftime"):
                row_vals[0] = row_vals[0].strftime("%d/%m/%Y")
            ws.append(row_vals)
            for cell in ws[ws.max_row]:
                _style_data_cell(cell)

        return ws.max_row + 2  # próxima tabla con espacio

    # Tabla semanal (columna A)
    end_row_weekly = write_table(1, "Tendencia Semanal — % Ingresos Oportunos", weekly_df)

    # Gráfico de tendencia semanal
    if not weekly_df.empty and len(weekly_df) > 1:
        data_start = 3  # fila donde comienzan los datos (1=título, 2=cabecera, 3=datos)
        data_end   = 2 + len(weekly_df)
        pct_col    = weekly_df.columns.tolist().index("% Oportuno") + 1  # 1-indexed

        chart = LineChart()
        chart.title  = "% Ingresos Oportunos por Semana"
        chart.y_axis.title = "% Oportuno"
        chart.x_axis.title = "Semana"
        chart.style  = 10
        chart.height = 12
        chart.width  = 22

        data_ref = Reference(ws, min_col=pct_col, min_row=2, max_row=data_end)
        cats_ref = Reference(ws, min_col=1,       min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.line.solidFill = "FF2F5597"
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size   = 5

        ws.add_chart(chart, f"F1")

    # Tabla mensual debajo de la semanal (o en columna E si hay espacio)
    write_table(end_row_weekly, "Tendencia Mensual — % Ingresos Oportunos", monthly_df)

    auto_column_width(ws)


# ── Hoja de recomendaciones ───────────────────────────────────────────────────

def write_recommendations_sheet(wb: Workbook, recommendations: list):
    ws = wb.create_sheet(OUTPUT_SHEETS["recommendations"])
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"].value = "RECOMENDACIONES BASADAS EN LOS HALLAZGOS"
    ws["A1"].fill  = _fill(COLOR_HEADER_ORANGE)
    ws["A1"].font  = _font(bold=True, color=COLOR_HEADER_FONT, size=13)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    if not recommendations:
        ws.merge_cells("A2:C2")
        ws["A2"].value = "No se generaron recomendaciones automáticas."
        return

    for i, rec in enumerate(recommendations, 1):
        row_idx = i + 1
        ws.merge_cells(
            start_row=row_idx, start_column=1,
            end_row=row_idx, end_column=3
        )
        cell = ws.cell(row_idx, 1)
        cell.value = f"{i}. {rec}"
        cell.font  = _font(size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True)
        cell.border = _border()
        bg = COLOR_ALT_ROW if i % 2 == 0 else None
        if bg:
            cell.fill = _fill(bg)
        ws.row_dimensions[row_idx].height = 50

    ws.column_dimensions["A"].width = MAX_COL_WIDTH
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10


# ── Guardar workbook ──────────────────────────────────────────────────────────

def save_workbook(wb: Workbook, output_path: str):
    """Guarda el workbook y activa la primera hoja."""
    wb.active = wb[OUTPUT_SHEETS["kpis"]]
    wb.save(output_path)
    print(f"\n  Excel generado exitosamente en:\n  {output_path}")
